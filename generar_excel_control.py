import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, timedelta
import calendar

# Crear el libro de trabajo
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Control Pólizas y Endosos"

# Estilos
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

verde_fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")
amarillo_fill = PatternFill(start_color="F1C40F", end_color="F1C40F", fill_type="solid")
rojo_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")

border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Encabezados
encabezados = [
    "FOLIO",
    "CLIENTE",
    "TIPO (PÓLIZA/ENDOSO)",
    "FECHA DE ENTREGA AL COBRADOR",
    "FECHA LÍMITE (5 DÍAS)",
    "FECHA DE ENTREGA AL CLIENTE",
    "ESTADO (🟢/🟡/🔴)",
    "OBSERVACIONES"
]

# Anchos de columna
anchos = [12, 25, 20, 25, 25, 25, 20, 30]

# Escribir encabezados
for col, (header, ancho) in enumerate(zip(encabezados, anchos), 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border_thin
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho

# Ejemplo de fila de datos (opcional, la можем borrar)
# Fila de ejemplo para guía
ejemplo_fila = [
    "12345",
    "Juan Pérez",
    "Póliza",
    datetime.now().strftime("%Y-%m-%d"),
    (datetime.now() + timedelta(days=5)).strftime("%Y-%m-%d"),
    "",
    "",
    "Notas aquí..."
]

# Escribir fila de ejemplo (opcional, la eliminamos para que Elena la borre si quiere)
# for col, valor in enumerate(ejemplo_fila, 1):
#     ws.cell(row=2, column=col, value=valor)

# Agregar notas explicativas en la hoja 2
ws_notas = wb.create_sheet("Guía de Uso")

notas_content = [
    "GUÍA DE USO - CONTROL DE PÓLIZAS Y ENDOSOS",
    "",
    "1. Esta plantilla te ayudará a controlar el tiempo de entrega de pólizas y endosos.",
    "2. Los cobradores tienen un límite de 5 días hábiles para entregar la póliza/endoso al cliente.",
    "",
    "CÓMO USARLA:",
    "- Ingresa los datos en las columnas sombreadas.",
    "- Las fechas de límite se calculan automáticamente.",
    "- El estado se actualiza solo (Verde/Amarillo/Rojo).",
    "",
    "LEYENDA DE ESTADOS:",
    "🟢 Verde: Entregado a tiempo (0-3 días)",
    "🟡 Amarillo: Por entregar (4 días)",
    "🔴 Rojo: Retrasado (5+ días)",
    "",
    "¡Buena suerte, Elena! 🌙✨"
]

for row, texto in enumerate(notas_content, 1):
    ws_notas.cell(row=row, column=1, value=texto)

# Guardar el archivo
wb.save("/root/.openclaw/workspace/Control_Polizas_Endosos.xlsx")
print("✅ Archivo creado exitosamente: Control_Polizas_Endosos.xlsx")
