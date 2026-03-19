#!/usr/bin/env python3
"""Resumen de cobranza desde Excel del sistema.

- Filtra filas con estatus ENTREGADO
- Reporta totales por día
- Detecta pagos de $125 (endosos) para exclusión
- Lista folios por día

Diseñado para el formato típico:
  hoja: Cobranza
  encabezados en fila 4
  columnas: FECHA, FOLIO, MONTO, ENTREGADO EN OFICINA

Si los encabezados varían, usar flags para ajustar.
"""

import argparse
import datetime as dt
import re
from collections import defaultdict

import openpyxl


def norm_header(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).strip()).upper()


def parse_date(x):
    if isinstance(x, dt.datetime):
        return x.date()
    if isinstance(x, dt.date):
        return x
    if isinstance(x, str):
        x = x.strip()
        m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", x)
        if m:
            d, mth, y = map(int, m.groups())
            return dt.date(y, mth, d)
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", required=True, help="Ruta al .xlsx")
    ap.add_argument("--sheet", default="Cobranza", help="Nombre de hoja")
    ap.add_argument("--header-row", type=int, default=4, help="Fila de encabezados (1-indexed)")

    ap.add_argument("--fecha-col", default="FECHA")
    ap.add_argument("--folio-col", default="FOLIO")
    ap.add_argument("--monto-col", default="MONTO")

    ap.add_argument("--entregado-col", default="ENTREGADO EN OFICINA")
    ap.add_argument("--entregado-value", default="ENTREGADO")

    ap.add_argument("--exclude-125", action="store_true", help="Excluir montos == 125")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.excel, data_only=True)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"Hoja '{args.sheet}' no existe. Hojas: {wb.sheetnames}")
    ws = wb[args.sheet]

    hdr = [c.value for c in ws[args.header_row]]
    idx = {norm_header(v): i for i, v in enumerate(hdr) if v is not None}

    def get_col(name):
        key = norm_header(name)
        if key not in idx:
            raise SystemExit(f"No encuentro columna '{name}'. Encabezados detectados: {list(idx.keys())}")
        return idx[key]

    fecha_i = get_col(args.fecha_col)
    folio_i = get_col(args.folio_col)
    monto_i = get_col(args.monto_col)
    est_i = get_col(args.entregado_col)

    entregado_value = norm_header(args.entregado_value)

    totals = defaultdict(float)
    folios = defaultdict(list)
    endosos_125 = []

    for row in ws.iter_rows(min_row=args.header_row + 1, values_only=True):
        fecha = parse_date(row[fecha_i])
        if not fecha:
            continue

        est = row[est_i]
        est_norm = norm_header(est) if est is not None else ""
        if est_norm != entregado_value:
            continue

        folio = row[folio_i]
        monto = row[monto_i]
        if monto is None:
            continue

        try:
            monto = float(monto)
        except Exception:
            continue

        if abs(monto - 125.0) < 1e-9:
            endosos_125.append((fecha, folio, monto))
            if args.exclude_125:
                continue

        totals[fecha] += monto
        folios[fecha].append((folio, monto))

    print("=== RESUMEN SISTEMA ===")
    print(f"Archivo: {args.excel}")
    print(f"Hoja: {args.sheet}")
    print(f"Filtro: {args.entregado_col} == {args.entregado_value}")
    print(f"Excluir $125: {'SI' if args.exclude_125 else 'NO'}")

    grand = sum(totals.values())
    print("\nTotales por día:")
    for d in sorted(totals.keys()):
        print(f"- {d.strftime('%d/%m/%Y')}: {totals[d]:,.0f}")

    print(f"\nTOTAL: {grand:,.0f}")

    print("\nPagos $125 detectados (para revisión/exclusión):")
    if not endosos_125:
        print("- (ninguno)")
    else:
        for d, fol, m in sorted(endosos_125, key=lambda x: (x[0], str(x[1]))):
            print(f"- {d.strftime('%d/%m/%Y')}: folio {fol} monto {m:,.0f}")

    print("\nFolios por día (folio, monto):")
    for d in sorted(folios.keys()):
        items = ", ".join([f"{f}:{m:,.0f}" for f, m in folios[d]])
        print(f"- {d.strftime('%d/%m/%Y')}: {items}")


if __name__ == "__main__":
    main()
