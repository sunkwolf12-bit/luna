"""Reconstruye la hoja DATOS GRAL. de un xlsx de cobranza.

FUENTE CANONICA: ``backend/scripts/reconstruir_datos_gral.py`` en el repo
``sunkwolf/sicc``. Esta copia vive aqui para que la skill ``sicc-ingesta``
(deploy en lunita) pueda ejecutarla como pre-parse step sin depender del
repo backend. Si tocas la logica, modifica primero el canonico y vuelve a
copiar aqui (`cp backend/scripts/reconstruir_datos_gral.py
skill/sicc-ingesta/scripts/` + restaurar este header). Tests viven en
``backend/tests/test_reconstruir_datos_gral.py`` (11 casos, suite
standalone sin DB).

Copia el total mensual (TOTAL MENSUAL, col B) de cada hoja ``COBRANZA *``
hacia la columna correspondiente de la hoja ``DATOS GRAL.``.

Problema que resuelve:
    Luna llenó los top5 individuales del xlsx 2024 (hojas ``COBRANZA *``)
    pero olvidó replicar los totales en la hoja resumen ``DATOS GRAL.``.
    Los meses quedaron con ``total=0`` al parsear. Este script corrige el
    xlsx antes de enviarlo a la ingesta, sin modificar las hojas ``COBRANZA``.

Mapeo concepto -> columna en DATOS GRAL. (header en R3, datos R4..R15):
    A=MES  B=CORRIENTE  C=%  D=CANCELACIONES  E=%  F=EFECTIVA  G=%
    H=RECUPERADA  I=%  J=ANTICIPADA FUTURA  K=%  L=VENCIDA  M=%
    N=ANTICIPADA ANT.  O=%  P=COBRANZA TOTAL

Hojas fuente (col A=MES, col B=TOTAL MENSUAL) con sus conceptos:
    COBRANZA CORRIENTE      -> col B
    CANCELACIONES           -> col D
    COBRANZA EFECTIVA       -> col F
    COBRANZA RECUPERADA     -> col H
    COBRANZA ADELANTADA F.  -> col J
    COBRANZA VENCIDA        -> col L
    (ANTICIPADA ANT. no tiene hoja fuente; se deja sin tocar)

Uso:
    # In-place (sobreescribe el xlsx original):
    python reconstruir_datos_gral.py --xlsx ESTADISTICA_2024.xlsx --inplace

    # Output separado (original intacto):
    python reconstruir_datos_gral.py --xlsx ESTADISTICA_2024.xlsx \\
        --out ESTADISTICA_2024_FIXED.xlsx

    # Forzar sobrescritura de celdas ya llenas:
    python reconstruir_datos_gral.py --xlsx ESTADISTICA_2024.xlsx \\
        --inplace --force

Sin emojis. Decimal nunca float. Type hints estrictos.
"""

from __future__ import annotations

import argparse
import logging
import shutil
import sys
from decimal import Decimal
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:  # pragma: no cover
    sys.stderr.write(
        "Falta openpyxl. Instalar con `pip install openpyxl`.\n"
    )
    sys.exit(2)

logger = logging.getLogger("sicc.reconstruir_datos_gral")

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

# Mapeo: concepto -> (nombre_hoja_fuente, columna_en_DATOS_GRAL)
# El orden importa para el logging; los conceptos sin hoja se dejan fuera.
MAPEO_CONCEPTOS: tuple[tuple[str, str, str], ...] = (
    ("CORRIENTE", "COBRANZA CORRIENTE", "B"),
    ("CANCELACIONES", "CANCELACIONES", "D"),
    ("EFECTIVA", "COBRANZA EFECTIVA", "F"),
    ("RECUPERADA", "COBRANZA RECUPERADA", "H"),
    ("ANTICIPADA_FUTURA", "COBRANZA ADELANTADA F.", "J"),
    ("VENCIDA", "COBRANZA VENCIDA", "L"),
)

# Meses en orden + fila correspondiente en DATOS GRAL. (header R3, datos R4..R15).
MESES_A_FILA: dict[str, int] = {
    "ENERO": 4,
    "FEBRERO": 5,
    "MARZO": 6,
    "ABRIL": 7,
    "MAYO": 8,
    "JUNIO": 9,
    "JULIO": 10,
    "AGOSTO": 11,
    "SEPTIEMBRE": 12,
    "OCTUBRE": 13,
    "NOVIEMBRE": 14,
    "DICIEMBRE": 15,
}

HOJA_GRAL = "DATOS GRAL."
# Fila minima de datos en las hojas COBRANZA (header en R3 o R4; datos desde R5).
COBRANZA_DATA_START = 5


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _normalizar_mes(valor: Any) -> str | None:
    """Convierte el valor de celda a clave de mes normalizada (mayusculas sin acento).

    Devuelve None si el valor no es un mes reconocible.
    """
    if valor is None:
        return None
    if not isinstance(valor, str):
        return None
    key = valor.strip().upper()
    # Normalizar acentos comunes en nombres de mes espanol.
    for con_acento, sin_acento in (
        ("Á", "A"), ("É", "E"), ("Í", "I"), ("Ó", "O"), ("Ú", "U"),
    ):
        key = key.replace(con_acento, sin_acento)
    return key if key in MESES_A_FILA else None


def _celda_tiene_valor(valor: Any) -> bool:
    """Devuelve True si la celda tiene un valor numerico no-cero."""
    if valor is None:
        return False
    if isinstance(valor, str):
        s = valor.strip()
        if not s:
            return False
        try:
            return Decimal(s.replace(",", "")) != 0
        except Exception:
            return False
    if isinstance(valor, (int, float)):
        return valor != 0
    if isinstance(valor, Decimal):
        return valor != 0
    return False


def _leer_total_cobranza(ws: Worksheet) -> dict[str, Decimal]:
    """Lee los pares (mes_key, total_mensual) de una hoja COBRANZA.

    La estructura esperada: col A = MES, col B = TOTAL MENSUAL.
    Datos a partir de la fila COBRANZA_DATA_START (saltea header).
    Ignora filas sin mes valido o total nulo/cero.
    """
    totales: dict[str, Decimal] = {}
    for r in range(COBRANZA_DATA_START, ws.max_row + 1):
        mes_raw = ws.cell(row=r, column=1).value
        total_raw = ws.cell(row=r, column=2).value
        mes_key = _normalizar_mes(mes_raw)
        if mes_key is None:
            continue
        if total_raw is None:
            continue
        if isinstance(total_raw, (int, float)):
            total_d = Decimal(str(total_raw))
        elif isinstance(total_raw, Decimal):
            total_d = total_raw
        elif isinstance(total_raw, str):
            s = total_raw.strip().replace(",", "")
            if not s:
                continue
            try:
                total_d = Decimal(s)
            except Exception:
                logger.warning(
                    "hoja %r fila %d: total no parseable %r; saltando",
                    ws.title, r, total_raw,
                )
                continue
        else:
            continue
        if total_d == 0:
            continue
        totales[mes_key] = total_d
    return totales


# ---------------------------------------------------------------------------
# Logica principal
# ---------------------------------------------------------------------------


def reconstruir(
    wb: Workbook,
    *,
    force: bool = False,
) -> dict[str, Any]:
    """Aplica la reconstruccion sobre el Workbook en memoria.

    Devuelve un dict con metricas:
        cambios: int              - celdas escritas
        saltadas_no_cero: int     - celdas con valor existente (no forzadas)
        advertencias: list[str]   - problemas no-fatales
    """
    if HOJA_GRAL not in wb.sheetnames:
        raise ValueError(
            f"xlsx no tiene hoja '{HOJA_GRAL}'. "
            f"Hojas presentes: {wb.sheetnames!r}"
        )
    ws_gral: Worksheet = wb[HOJA_GRAL]

    sheetnames_upper: dict[str, str] = {
        n.upper().strip(): n for n in wb.sheetnames
    }

    cambios = 0
    saltadas_no_cero = 0
    advertencias: list[str] = []

    for concepto, nombre_hoja, col_dest in MAPEO_CONCEPTOS:
        hoja_key = nombre_hoja.upper().strip()
        if hoja_key not in sheetnames_upper:
            msg = (
                f"concepto {concepto!r}: hoja fuente {nombre_hoja!r} "
                f"no encontrada en xlsx; saltando"
            )
            logger.warning("%s", msg)
            advertencias.append(msg)
            continue

        ws_src: Worksheet = wb[sheetnames_upper[hoja_key]]
        totales_hoja = _leer_total_cobranza(ws_src)

        if not totales_hoja:
            msg = (
                f"concepto {concepto!r}: hoja {nombre_hoja!r} no tiene "
                f"totales validos (vacia o todos cero)"
            )
            logger.warning("%s", msg)
            advertencias.append(msg)
            # No es error fatal; continuamos con los demas conceptos.

        for mes_key, total in totales_hoja.items():
            fila = MESES_A_FILA[mes_key]
            coord = f"{col_dest}{fila}"
            valor_actual = ws_gral[coord].value

            if _celda_tiene_valor(valor_actual):
                if not force:
                    logger.debug(
                        "DATOS GRAL. %s ya tiene valor %r; saltando "
                        "(pasa --force para sobreescribir)",
                        coord, valor_actual,
                    )
                    saltadas_no_cero += 1
                    continue
                else:
                    logger.info(
                        "DATOS GRAL. %s: valor existente %r -> %s [forzado]",
                        coord, valor_actual, total,
                    )
            else:
                logger.info(
                    "DATOS GRAL. %s: vacio -> %s  [%s / %s]",
                    coord, total, mes_key, concepto,
                )

            # openpyxl espera float o int para celdas numericas; usamos float
            # SOLO para la escritura en xlsx (la logica interna opera con Decimal).
            ws_gral[coord] = float(total)
            cambios += 1

        # Advertencia de coherencia inversa: DATOS GRAL. tiene valor pero la
        # hoja COBRANZA no tiene dato. No rompe; es informativo.
        for mes_key, fila in MESES_A_FILA.items():
            coord = f"{col_dest}{fila}"
            valor_gral = ws_gral[coord].value
            if _celda_tiene_valor(valor_gral) and mes_key not in totales_hoja:
                msg = (
                    f"coherencia: DATOS GRAL. {coord} tiene valor {valor_gral!r} "
                    f"pero {nombre_hoja!r} no tiene total para {mes_key}. "
                    f"Verificar el xlsx."
                )
                logger.warning("%s", msg)
                advertencias.append(msg)

    logger.info(
        "reconstruccion completa: %d celdas escritas, "
        "%d saltadas (no-cero sin --force), %d advertencias",
        cambios, saltadas_no_cero, len(advertencias),
    )
    return {
        "cambios": cambios,
        "saltadas_no_cero": saltadas_no_cero,
        "advertencias": advertencias,
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def _parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description=(
            "Reconstruye la hoja DATOS GRAL. copiando totales desde "
            "hojas COBRANZA *. Util cuando Luna lleno top5 individuales "
            "pero olvido replicar los totales en el resumen."
        ),
    )
    src_group = p.add_mutually_exclusive_group()
    src_group.add_argument(
        "--inplace",
        action="store_true",
        default=False,
        help=(
            "Sobreescribe el xlsx original. Mutuamente exclusivo con --out. "
            "Se hace una copia de seguridad antes de guardar."
        ),
    )
    src_group.add_argument(
        "--out",
        type=Path,
        default=None,
        metavar="OUTPUT",
        help=(
            "Ruta del xlsx de salida. El archivo original queda intacto. "
            "Mutuamente exclusivo con --inplace."
        ),
    )
    p.add_argument(
        "--xlsx",
        type=Path,
        required=True,
        help="Ruta al xlsx de cobranza a procesar.",
    )
    p.add_argument(
        "--force",
        action="store_true",
        default=False,
        help=(
            "Sobreescribe celdas de DATOS GRAL. que ya tengan valor no-cero. "
            "Por defecto, las celdas llenas se saltan (idempotencia)."
        ),
    )
    p.add_argument(
        "--log-level",
        type=str,
        default="INFO",
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
        help="Nivel de logging (default INFO).",
    )
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv)

    logging.basicConfig(
        level=getattr(logging, args.log_level.upper(), logging.INFO),
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )

    if not args.xlsx.exists():
        logger.error("xlsx no existe: %s", args.xlsx)
        return 2

    if not args.inplace and args.out is None:
        logger.error(
            "Se requiere --inplace o --out OUTPUT. "
            "Ver --help para mas detalle."
        )
        return 2

    # Definir ruta de destino.
    xlsx_path: Path = args.xlsx.resolve()
    if args.inplace:
        out_path: Path = xlsx_path
        # Copia de seguridad antes de sobreescribir.
        backup_path = xlsx_path.with_suffix(".bak.xlsx")
        shutil.copy2(xlsx_path, backup_path)
        logger.info("copia de seguridad guardada en %s", backup_path)
    else:
        assert args.out is not None
        out_path = args.out.resolve()

    logger.info("cargando xlsx: %s", xlsx_path)
    wb = load_workbook(str(xlsx_path))

    try:
        resultado = reconstruir(wb, force=args.force)
    except ValueError as exc:
        logger.error("error de estructura xlsx: %s", exc)
        return 1

    logger.info("guardando en: %s", out_path)
    wb.save(str(out_path))

    logger.info(
        "listo. cambios=%d saltadas=%d advertencias=%d",
        resultado["cambios"],
        resultado["saltadas_no_cero"],
        len(resultado["advertencias"]),
    )
    for adv in resultado["advertencias"]:
        logger.warning("  ADV: %s", adv)

    return 0


if __name__ == "__main__":  # pragma: no cover
    sys.exit(main())
