#!/usr/bin/env python
from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


PAGE_SIZE = landscape(letter)
PAGE_W, PAGE_H = PAGE_SIZE

NAVY = colors.HexColor("#0E2A3B")
TEAL = colors.HexColor("#0F766E")
GOLD = colors.HexColor("#F3B23C")
INK = colors.HexColor("#17212B")
MUTED = colors.HexColor("#667085")
LINE = colors.HexColor("#D7DEE8")
SOFT = colors.HexColor("#F6F8FB")
SOFT_TEAL = colors.HexColor("#EEF8F6")


@dataclass
class Section:
    sheet: str
    title: str
    start_row: int
    end_row: int
    min_col: int
    max_col: int


def register_fonts() -> tuple[str, str]:
    font_candidates = [
        # Windows
        (Path(r"C:\Windows\Fonts\calibri.ttf"), Path(r"C:\Windows\Fonts\calibrib.ttf"), "Calibri", "Calibri-Bold"),
        (Path(r"C:\Windows\Fonts\arial.ttf"), Path(r"C:\Windows\Fonts\arialbd.ttf"), "Arial", "Arial-Bold"),
        # Linux (DejaVu / Liberation)
        (Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"), Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"), "DejaVuSans", "DejaVuSans-Bold"),
        (Path("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"), Path("/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"), "LiberationSans", "LiberationSans-Bold"),
        # macOS
        (Path("/Library/Fonts/Arial.ttf"), Path("/Library/Fonts/Arial Bold.ttf"), "Arial", "Arial-Bold"),
    ]
    for regular, bold, regular_name, bold_name in font_candidates:
        if regular.exists() and bold.exists():
            pdfmetrics.registerFont(TTFont(regular_name, str(regular)))
            pdfmetrics.registerFont(TTFont(bold_name, str(bold)))
            return regular_name, bold_name
    return "Helvetica", "Helvetica-Bold"


def clean_text(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    text = text.strip('"')
    text = re.sub(r"\s+", " ", text)
    return text


def is_number(value) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def compact_money(value: float) -> str:
    sign = "-" if value < 0 else ""
    value = abs(float(value))
    if value >= 1_000_000:
        return f"{sign}${value / 1_000_000:,.1f}M".replace(".0M", "M")
    if value >= 1_000:
        return f"{sign}${value / 1_000:,.1f}K".replace(".0K", "K")
    return f"{sign}${value:,.0f}"


def fmt_value(value, row_label: str = "") -> str:
    if value is None:
        return ""
    if is_number(value):
        label = row_label.lower()
        if ("%" in label or "percent" in label or "dismin" in label or "rate" in label) and abs(value) <= 1:
            return f"{value:.1%}"
        return compact_money(value)
    return clean_text(value)


def nonempty_cells(ws, row: int, min_col: int = 1, max_col: int | None = None) -> list[tuple[int, object]]:
    max_col = max_col or ws.max_column
    cells = []
    for col in range(min_col, max_col + 1):
        value = ws.cell(row, col).value
        if value is not None and clean_text(value) != "":
            cells.append((col, value))
    return cells


def row_is_title(cells: list[tuple[int, object]]) -> bool:
    if not cells:
        return False
    text_cells = [clean_text(v) for _, v in cells if isinstance(v, str) and clean_text(v)]
    if len(cells) <= 2 and text_cells:
        text = " ".join(text_cells)
        upper = text.upper()
        if any(key in upper for key in ["EJERCICIO", "PROJECT", "PROYECTO", "ACTIVIDAD", "SECTION", "TEMA"]):
            return True
        if len(text) > 8 and upper == text and sum(ch.isalpha() for ch in text) >= 5:
            return True
    return False


def find_sections(ws) -> list[Section]:
    used_rows = [r for r in range(1, ws.max_row + 1) if nonempty_cells(ws, r)]
    if not used_rows:
        return []

    title_rows = []
    for row in used_rows:
        cells = nonempty_cells(ws, row)
        if row_is_title(cells):
            title_rows.append(row)

    if not title_rows:
        title = ws.title
        min_col, max_col = used_col_bounds(ws, min(used_rows), max(used_rows))
        return [Section(ws.title, title, min(used_rows), max(used_rows), min_col, max_col)]

    sections = []
    for index, title_row in enumerate(title_rows):
        next_title = title_rows[index + 1] if index + 1 < len(title_rows) else max(used_rows) + 1
        end_row = next_title - 1
        while end_row > title_row and not nonempty_cells(ws, end_row):
            end_row -= 1
        min_col, max_col = used_col_bounds(ws, title_row, end_row)
        title = clean_text(ws.cell(title_row, min_col).value) or ws.title
        sections.append(Section(ws.title, title, title_row, end_row, min_col, max_col))
    return sections


def used_col_bounds(ws, start_row: int, end_row: int) -> tuple[int, int]:
    cols = []
    for row in range(start_row, end_row + 1):
        cols.extend(col for col, _ in nonempty_cells(ws, row))
    return (min(cols), max(cols)) if cols else (1, 1)


def find_main_note_split(ws, section: Section) -> tuple[int, int]:
    empty_counts = {}
    row_count = max(1, section.end_row - section.start_row + 1)
    for col in range(section.min_col + 1, section.max_col):
        empty = 0
        for row in range(section.start_row, section.end_row + 1):
            if clean_text(ws.cell(row, col).value) == "":
                empty += 1
        empty_counts[col] = empty / row_count

    candidates = [col for col, ratio in empty_counts.items() if ratio >= 0.65]
    if not candidates:
        return section.max_col, section.max_col + 1

    split_col = max(candidates, key=lambda c: (empty_counts[c], c))
    left_nonempty = sum(1 for r in range(section.start_row, section.end_row + 1) for c in range(section.min_col, split_col) if clean_text(ws.cell(r, c).value))
    right_nonempty = sum(1 for r in range(section.start_row, section.end_row + 1) for c in range(split_col + 1, section.max_col + 1) if clean_text(ws.cell(r, c).value))
    if left_nonempty >= 4 and right_nonempty >= 2:
        return split_col - 1, split_col + 1
    return section.max_col, section.max_col + 1


def find_table_rows(ws, section: Section, main_end_col: int) -> tuple[int, int]:
    rows = []
    for row in range(section.start_row, section.end_row + 1):
        count = len(nonempty_cells(ws, row, section.min_col, main_end_col))
        if count >= 2:
            rows.append(row)
    if not rows:
        return section.start_row, section.end_row
    return min(rows), max(rows)


def collect_notes(ws, section: Section, note_start_col: int) -> list[str]:
    if note_start_col > section.max_col:
        return []
    notes = []
    for row in range(section.start_row, section.end_row + 1):
        for col in range(note_start_col, section.max_col + 1):
            text = clean_text(ws.cell(row, col).value)
            if text:
                notes.append(text)
    return notes


def split_notes(notes: Iterable[str]) -> tuple[list[str], list[str]]:
    formulas = []
    interpretation = []
    mode = "formulas"
    for note in notes:
        upper = note.upper()
        if "INTERPRET" in upper or "CONCLUSION" in upper or "CONCLUSI" in upper:
            mode = "interpretation"
            continue
        if "FORMULA" in upper or "FÓRMULA" in upper:
            continue
        if mode == "formulas":
            formulas.append(note)
        else:
            interpretation.append(note)
    return formulas, interpretation


def make_styles(font_regular: str, font_bold: str):
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="CoverKicker", fontName=font_bold, fontSize=13, leading=16, textColor=GOLD, alignment=TA_CENTER))
    styles.add(ParagraphStyle(name="CoverTitle", fontName=font_bold, fontSize=30, leading=34, textColor=colors.white, alignment=TA_CENTER, spaceAfter=12))
    styles.add(ParagraphStyle(name="CoverMeta", fontName=font_regular, fontSize=12, leading=15, textColor=colors.HexColor("#D7E8EF"), alignment=TA_CENTER))
    styles.add(ParagraphStyle(name="SectionTitle", fontName=font_bold, fontSize=16, leading=19, textColor=NAVY))
    styles.add(ParagraphStyle(name="Cell", fontName=font_regular, fontSize=7.2, leading=8.4, textColor=INK, alignment=TA_RIGHT))
    styles.add(ParagraphStyle(name="CellLeft", fontName=font_regular, fontSize=7.2, leading=8.4, textColor=INK, alignment=TA_LEFT))
    styles.add(ParagraphStyle(name="CellHeader", fontName=font_bold, fontSize=7.5, leading=8.5, textColor=colors.white, alignment=TA_CENTER))
    styles.add(ParagraphStyle(name="BoxTitle", fontName=font_bold, fontSize=10, leading=12, textColor=NAVY, spaceAfter=5))
    styles.add(ParagraphStyle(name="BoxText", fontName=font_regular, fontSize=8.2, leading=10.2, textColor=INK))
    return styles


def para(text: str, style) -> Paragraph:
    return Paragraph(clean_text(text), style)


def make_eyebrow(title: str, accent, styles, width: float) -> Table:
    table = Table([["", para(title, styles["SectionTitle"])]], colWidths=[0.12 * inch, width - 0.12 * inch])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), accent),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    return table


def make_box(title: str, items: list[str], styles, width: float, fill) -> Table:
    rows = [[para(title, styles["BoxTitle"])]]
    rows.extend([[para(item, styles["BoxText"])] for item in items[:18]])
    if len(items) > 18:
        rows.append([para(f"+ {len(items) - 18} notas adicionales omitidas por espacio; revisar fuente.", styles["BoxText"])])
    if len(rows) == 1:
        rows.append([para("Sin notas adicionales.", styles["BoxText"])])
    table = Table(rows, colWidths=[width], hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), fill),
        ("BOX", (0, 0), (-1, -1), 0.8, LINE),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    return table


def table_data(ws, section: Section, start_row: int, end_row: int, end_col: int, styles) -> list[list[Paragraph]]:
    rows = []
    for row in range(start_row, end_row + 1):
        row_label = clean_text(ws.cell(row, section.min_col).value)
        output_row = []
        for col in range(section.min_col, end_col + 1):
            style = styles["CellHeader"] if row == start_row else (styles["CellLeft"] if col == section.min_col else styles["Cell"])
            output_row.append(para(fmt_value(ws.cell(row, col).value, row_label), style))
        rows.append(output_row)
    return rows


def make_data_table(data: list[list[Paragraph]], col_widths: list[float]) -> Table:
    table = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("BOX", (0, 0), (-1, -1), 0.8, LINE),
        ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#E3E8EF")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for row_index in range(1, len(data)):
        if row_index % 2:
            style.append(("BACKGROUND", (0, row_index), (-1, row_index), SOFT))
        label = clean_text(data[row_index][0].getPlainText()).lower()
        if any(key in label for key in ["uai", "uaii", "ebit", "total", "utilidad", "net", "resultado", "balance"]):
            style.append(("BACKGROUND", (0, row_index), (-1, row_index), SOFT_TEAL))
    table.setStyle(TableStyle(style))
    return table


def section_flowables(ws, section: Section, styles, accent) -> list:
    available_w = PAGE_W - 0.8 * inch
    main_end_col, note_start_col = find_main_note_split(ws, section)
    table_start, table_end = find_table_rows(ws, section, main_end_col)
    data = table_data(ws, section, table_start, table_end, main_end_col, styles)
    cols = max(1, main_end_col - section.min_col + 1)
    note_w = 2.4 * inch if note_start_col <= section.max_col else 0
    gutter = 0.16 * inch if note_w else 0
    table_w = available_w - note_w - gutter
    first_col = min(1.8 * inch, table_w * 0.34)
    other_col = (table_w - first_col) / max(1, cols - 1)
    col_widths = [first_col] + [other_col] * (cols - 1)
    main_table = make_data_table(data, col_widths)
    flowables = [make_eyebrow(section.title, accent, styles, available_w), Spacer(1, 8)]

    notes = collect_notes(ws, section, note_start_col)
    formulas, interpretation = split_notes(notes)

    if note_w:
        side = [
            make_box("Formulas clave", formulas, styles, note_w, colors.HexColor("#FFF8EA")),
            Spacer(1, 9),
            make_box("Interpretacion", interpretation, styles, note_w, colors.HexColor("#EFF7FF")),
        ]
        layout = Table([[main_table, side]], colWidths=[table_w, note_w], hAlign="LEFT")
        layout.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ]))
        flowables.append(layout)
    else:
        flowables.append(main_table)
    return flowables


def cover_page(canvas, doc):
    canvas.saveState()
    canvas.setFillColor(NAVY)
    canvas.rect(0, 0, PAGE_W, PAGE_H, fill=1, stroke=0)
    canvas.setFillColor(colors.HexColor("#13394F"))
    canvas.rect(0, 0, PAGE_W, 1.25 * inch, fill=1, stroke=0)
    canvas.setFillColor(TEAL)
    canvas.rect(0, PAGE_H - 0.18 * inch, PAGE_W, 0.18 * inch, fill=1, stroke=0)
    canvas.restoreState()


def page_frame(canvas, doc):
    canvas.saveState()
    canvas.setFillColor(colors.HexColor("#F8FAFC"))
    canvas.rect(0, 0, PAGE_W, PAGE_H, fill=1, stroke=0)
    canvas.setFillColor(NAVY)
    canvas.rect(0, PAGE_H - 0.34 * inch, PAGE_W, 0.34 * inch, fill=1, stroke=0)
    canvas.setFillColor(TEAL)
    canvas.rect(0, PAGE_H - 0.39 * inch, PAGE_W, 0.05 * inch, fill=1, stroke=0)
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(MUTED)
    canvas.drawRightString(PAGE_W - 0.42 * inch, 0.24 * inch, f"Page {doc.page}")
    canvas.restoreState()


def build_pdf(input_path: Path, output_path: Path, title: str | None = None):
    font_regular, font_bold = register_fonts()
    styles = make_styles(font_regular, font_bold)
    wb = load_workbook(input_path, data_only=True)
    sections = []
    for ws in wb.worksheets:
        sections.extend(find_sections(ws))
    if not sections:
        raise ValueError("No non-empty worksheet content found.")

    report_title = title or input_path.stem
    output_path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=PAGE_SIZE,
        leftMargin=0.4 * inch,
        rightMargin=0.4 * inch,
        topMargin=0.52 * inch,
        bottomMargin=0.45 * inch,
        title=report_title,
    )

    story = [
        Spacer(1, 1.45 * inch),
        para("Spreadsheet report", styles["CoverKicker"]),
        para(report_title, styles["CoverTitle"]),
        para(f"{len(sections)} section(s) from {len(wb.worksheets)} sheet(s)", styles["CoverMeta"]),
        Spacer(1, 0.55 * inch),
    ]
    cover_rows = [[para(section.title, styles["CoverMeta"])] for section in sections[:10]]
    cover_table = Table(cover_rows, colWidths=[5.8 * inch], hAlign="CENTER")
    cover_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#173D53")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#35596D")),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.extend([cover_table, PageBreak()])

    accents = [TEAL, colors.HexColor("#1D4ED8"), colors.HexColor("#7C3AED"), colors.HexColor("#BE123C"), colors.HexColor("#EA580C")]
    sheet_by_name = {ws.title: ws for ws in wb.worksheets}
    for idx, section in enumerate(sections):
        story.extend(section_flowables(sheet_by_name[section.sheet], section, styles, accents[idx % len(accents)]))
        if idx != len(sections) - 1:
            story.append(PageBreak())

    doc.build(story, onFirstPage=cover_page, onLaterPages=page_frame)


def parse_args():
    parser = argparse.ArgumentParser(description="Create a polished landscape PDF report from an Excel workbook.")
    parser.add_argument("--input", required=True, type=Path, help="Path to .xlsx workbook.")
    parser.add_argument("--output", required=True, type=Path, help="Path for the PDF output.")
    parser.add_argument("--title", default=None, help="Optional report title.")
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    build_pdf(args.input, args.output, args.title)
    print(args.output)
