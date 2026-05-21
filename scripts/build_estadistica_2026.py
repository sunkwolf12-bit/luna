"""
Crea plantilla profesional ESTADISTICA_PARA_JUNTA_MENSUAL_2026
Colores: Azul PROTEG-RT (#1F3C6E) + Dorado (#C9A84C)
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
import os

# ─── COLORES ───────────────────────────────────────────
AZUL_OBSCURO   = "1F3C6E"
AZUL_MEDIO     = "2E5090"
AZUL_CLARO     = "D6E4F0"
DORADO         = "C9A84C"
DORADO_CLARO   = "F5E6B8"
BLANCO         = "FFFFFF"
GRIS_CLARO     = "F2F2F2"
GRIS_TEXTO     = "404040"
NEGRO          = "1A1A1A"

# ─── BORDES ────────────────────────────────────────────
bordo = Border(
    left=Side(style='thin', color=AZUL_OBSCURO),
    right=Side(style='thin', color=AZUL_OBSCURO),
    top=Side(style='thin', color=AZUL_OBSCURO),
    bottom=Side(style='thin', color=AZUL_OBSCURO),
)
bordo_ext = Border(
    left=Side(style='medium', color=AZUL_OBSCURO),
    right=Side(style='medium', color=AZUL_OBSCURO),
    top=Side(style='medium', color=AZUL_OBSCURO),
    bottom=Side(style='medium', color=AZUL_OBSCURO),
)
bordo_dorado = Border(
    left=Side(style='thin', color=DORADO),
    right=Side(style='thin', color=DORADO),
    top=Side(style='thin', color=DORADO),
    bottom=Side(style='thin', color=DORADO),
)
thin = Side(style='hair', color=GRIS_TEXTO)
bordo_interno = Border(
    left=Side(style='hair', color=GRIS_TEXTO),
    right=Side(style='hair', color=GRIS_TEXTO),
    top=Side(style='hair', color=GRIS_TEXTO),
    bottom=Side(style='hair', color=GRIS_TEXTO),
)

# ─── HELPERS ────────────────────────────────────────────
def hdr(ws, row, col, text, 
        bg=AZUL_OBSCURO, fg=BLANCO, size=11, bold=True,
        merge_to=None, align="center", wrap=False):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Calibri", size=size, bold=bold, color=fg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                              wrap_text=wrap)
    if merge_to:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=merge_to)
    return cell

def cell_style(ws, row, col, value=None, 
               bg=BLANCO, fg=GRIS_TEXTO, bold=False, size=10,
               align="center", fmt=None, bord=None, wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Calibri", size=size, bold=bold, color=fg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                              wrap_text=wrap)
    if fmt:
        cell.number_format = fmt
    if bord:
        cell.border = bord
    return cell

def thin_border(ws, row, col):
    ws.cell(row=row, column=col).border = bordo_interno

def sección(ws, row, titulo, cols_cover, color=AZUL_MEDIO):
    """Barra de título de sección"""
    cell = ws.cell(row=row, column=1, value=titulo.upper())
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = Font(name="Calibri", size=10, bold=True, color=DORADO)
    cell.alignment = Alignment(horizontal="left", vertical="center",
                               indent=1)
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=cols_cover)
    # línea dorada abajo
    for c in range(1, cols_cover + 1):
        lc = ws.cell(row=row, column=c)
        lc.border = Border(bottom=Side(style='medium', color=DORADO))

# ─── DATOS DE MESES 2026 ────────────────────────────────
MESES_2026 = [
    ("ENERO",     "01"),
    ("FEBRERO",   "02"),
    ("MARZO",     "03"),
    ("ABRIL",     "04"),
    ("MAYO",      "05"),
    ("JUNIO",     "06"),
    ("JULIO",     "07"),
    ("AGOSTO",    "08"),
    ("SEPTIEMBRE","09"),
    ("OCTUBRE",   "10"),
    ("NOVIEMBRE", "11"),
    ("DICIEMBRE", "12"),
]

# Columnas que usa la hoja
# Col:  1        2         3        4        5        6        7        8        9       10       11       12        13
#       A        B         C        D        E        F        G        H        I        J        K        L        M        N
#       CONCEPTO | TOTALES | TOP-1   | TOP-2  | TOP-3  | TOP-4  | TOP-5  | %_TOP1 | %_TOP2 | %_TOP3 | %_TOP4 | %_TOP5 | %_ACUM
COLS = 14

# ─── CONSTRUCCIÓN ───────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "ESTADISTICA 2026"

# ══ 1) ENCABEZADO MAESTRO ════════════════════════════════
hdr(ws, 1, 1, "PROTEG-RT MUTUALIDAD",
    bg=AZUL_OBSCURO, fg=DORADO, size=16, bold=True,
    merge_to=COLS, align="center")
hdr(ws, 2, 1, "REPORTE DE COBRANZA — JUNTA MENSUAL",
    bg=AZUL_MEDIO, fg=BLANCO, size=12, bold=True,
    merge_to=COLS, align="center")
hdr(ws, 3, 1, "AÑO 2026",
    bg=AZUL_MEDIO, fg=DORADO, size=11, bold=False,
    merge_to=COLS, align="center")

# Fila vacía
ws.row_dimensions[4].height = 8

# ══ 2) COLUMNA DE ETIQUETAS (fija) ════════════════════════
ANCHO_COL_A = 32
ws.column_dimensions['A'].width = ANCHO_COL_A

# Anchos homogéneos para B-N
for c in range(2, COLS + 1):
    ws.column_dimensions[get_column_letter(c)].width = 13

# Fijar la columna A como fondo para facilitar scroll horizontal
ws.freeze_panes = "B4"

# ══ 3) FUNCIÓN: construir un mes ════════════════════════════
def build_month(ws, mes_nombre, mes_num, start_row):
    r = start_row

    # ── Encabezado del mes ────────────────────────────────
    hdr(ws, r, 1, f"{mes_nombre} 2026",
        bg=AZUL_OBSCURO, fg=DORADO, size=12, bold=True,
        merge_to=COLS, align="center")
    ws.row_dimensions[r].height = 22
    r += 1

    # ── Sub-encabezados de columnas ─────────────────────
    hdrs = ["CONCEPTO", "TOTAL ($)", "TOP 1", "TOP 2", "TOP 3",
            "TOP 4", "TOP 5", "% TOP 1", "% TOP 2", "% TOP 3",
            "% TOP 4", "% TOP 5", "% ACUM."]
    for c, h in enumerate(hdrs, 1):
        cell_style(ws, r, c, h,
                  bg=DORADO, fg=NEGRO, bold=True, size=9,
                  align="center")
    ws.row_dimensions[r].height = 28
    r += 1

    # helper para dato de categoría
    def cat_row(label, bg_cat=GRIS_CLARO):
        nonlocal r
        # Etiqueta
        cell_style(ws, r, 1, label,
                   bg=bg_cat, fg=GRIS_TEXTO, bold=True, size=10,
                   align="left")
        for c in range(2, COLS + 1):
            cell_style(ws, r, c, None,
                       bg=bg_cat, fg=GRIS_TEXTO, size=10,
                       align="center")
        r += 1

    def top5_row():
        nonlocal r
        for c in range(1, COLS + 1):
            cell_style(ws, r, c, None, bg=BLANCO, fg=GRIS_TEXTO,
                       size=10, align="center")
        r += 1

    # ── DATOS GENERALES ─────────────────────────────────
    sección(ws, r, "DATOS GENERALES", COLS, color=AZUL_MEDIO)
    r += 1
    cell_style(ws, r, 1, "Mes", bg=AZUL_CLARO, fg=GRIS_TEXTO,
               bold=True, align="left")
    cell_style(ws, r, 2, mes_nombre, bg=AZUL_CLARO, fg=NEGRO, bold=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    cell_style(ws, r, 5, "Año", bg=AZUL_CLARO, fg=GRIS_TEXTO,
               bold=True, align="left")
    cell_style(ws, r, 6, "2026", bg=AZUL_CLARO, fg=NEGRO, bold=True)
    r += 1
    cell_style(ws, r, 1, "Fecha de corte", bg=AZUL_CLARO, fg=GRIS_TEXTO,
               bold=True, align="left")
    cell_style(ws, r, 2, None, bg=AZUL_CLARO, fg=GRIS_TEXTO)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    cell_style(ws, r, 5, "Elaboró", bg=AZUL_CLARO, fg=GRIS_TEXTO,
               bold=True, align="left")
    cell_style(ws, r, 6, None, bg=AZUL_CLARO, fg=GRIS_TEXTO)
    r += 1
    cell_style(ws, r, 1, "Reportado por", bg=AZUL_CLARO, fg=GRIS_TEXTO,
               bold=True, align="left")
    cell_style(ws, r, 2, None, bg=AZUL_CLARO, fg=GRIS_TEXTO)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    cell_style(ws, r, 5, "Cargo", bg=AZUL_CLARO, fg=GRIS_TEXTO,
               bold=True, align="left")
    cell_style(ws, r, 6, None, bg=AZUL_CLARO, fg=GRIS_TEXTO)
    r += 1
    ws.row_dimensions[r].height = 6
    r += 1

    # ── CATEGORÍAS ───────────────────────────────────────
    cats = [
        ("COBRANZA CORRIENTE",        AZUL_CLARO),
        ("CANCELACIONES",            GRIS_CLARO),
        ("COBRANZA VENCIDA",          AZUL_CLARO),
        ("COBRANZA EFECTIVA",         GRIS_CLARO),
        ("COBRANZA RECUPERADA",       AZUL_CLARO),
        ("COBRANZA TOTAL GENERAL",    GRIS_CLARO),
        ("COBRANZA ADELANTADA F.",    AZUL_CLARO),
    ]
    for nom, bg in cats:
        sección(ws, r, nom, COLS, color=AZUL_MEDIO)
        r += 1
        # Fila TOTAL
        cell_style(ws, r, 1, "Total",
                   bg=bg, fg=GRIS_TEXTO, bold=True, align="left")
        cell_style(ws, r, 2, None,
                   bg=bg, fg=NEGRO, bold=True, size=11,
                   align="center", fmt='"$"#,##0.00')
        ws.row_dimensions[r].height = 20
        r += 1
        # Fila TOP 5 encabezado
        cell_style(ws, r, 1, "Top 5",
                   bg=DORADO_CLARO, fg=GRIS_TEXTO, bold=True,
                   size=9, align="center")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        for c in range(3, COLS + 1):
            cell_style(ws, r, c, None,
                       bg=DORADO_CLARO, fg=GRIS_TEXTO, size=9,
                       align="center")
        ws.row_dimensions[r].height = 18
        r += 1
        # Filas top 1-5
        top_labels = ["1er lugar", "2o lugar", "3er lugar",
                     "4o lugar", "5o lugar"]
        for tl in top_labels:
            cell_style(ws, r, 1, tl, bg=BLANCO, fg=GRIS_TEXTO,
                       size=9, align="left")
            for c in range(2, COLS + 1):
                cell_style(ws, r, c, None, bg=BLANCO, fg=GRIS_TEXTO,
                           size=9, align="center")
            ws.row_dimensions[r].height = 18
            r += 1
        ws.row_dimensions[r].height = 6
        r += 1

    # ── PROYECCIÓN ───────────────────────────────────────
    sección(ws, r, f"PROYECCIÓN {mes_nombre} → {mes_nombre.upper()}",
            COLS, color=DORADO)
    r += 1
    cell_style(ws, r, 1, f"Proyección {mes_nombre}",
              bg=DORADO_CLARO, fg=GRIS_TEXTO, bold=True, align="left")
    cell_style(ws, r, 2, None, bg=DORADO_CLARO, fg=NEGRO, bold=True,
               size=11, align="center", fmt='"$"#,##0.00')
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 2

    return r  # siguiente fila disponible

# ─── 4) CONSTRUIR TODOS LOS MESES ────────────────────────
r = 5  # Empezamos después del encabezado maestro
for mes_nombre, mes_num in MESES_2026:
    r = build_month(ws, mes_nombre, mes_num, r)
    ws.row_dimensions[r].height = 14  # separador entre meses
    r += 1

# ══ 5) RESUMEN ANUAL ═════════════════════════════════════
r += 2
hdr(ws, r, 1, "RESUMEN ANUAL 2026",
    bg=AZUL_OBSCURO, fg=DORADO, size=13, bold=True,
    merge_to=COLS, align="center")
ws.row_dimensions[r].height = 24
r += 1

# Tabla resumen 12 meses
hdrs_res = ["CONCEPTO"] + [m[0][:3].upper() for m in MESES_2026] + ["TOTAL", "% PART."]
for c, h in enumerate(hdrs_res, 1):
    cell_style(ws, r, c, h,
              bg=AZUL_MEDIO, fg=BLANCO, bold=True, size=9,
              align="center")
ws.row_dimensions[r].height = 26
r += 1

cats_resum = ["COBRANZA CORRIENTE", "CANCELACIONES", "COBRANZA VENCIDA",
              "COBRANZA EFECTIVA", "COBRANZA RECUPERADA",
              "COBRANZA TOTAL GENERAL", "COBRANZA ADELANTADA F.",
              "PROYECCIÓN"]
for i, nom in enumerate(cats_resum):
    bg_row = AZUL_CLARO if i % 2 == 0 else GRIS_CLARO
    cell_style(ws, r, 1, nom, bg=bg_row, fg=GRIS_TEXTO,
               bold=True, size=9, align="left")
    for c in range(2, 16):  # 12 meses + total + %
        cell_style(ws, r, c, None, bg=bg_row, fg=GRIS_TEXTO,
                   size=9, align="center")
    ws.row_dimensions[r].height = 18
    r += 1

# ══ 6) NOTAS / LEYENDA ══════════════════════════════════════
r += 2
hdr(ws, r, 1, "NOTAS Y LEYENDA",
    bg=AZUL_OBSCURO, fg=DORADO, size=10, bold=True,
    merge_to=COLS, align="center")
ws.row_dimensions[r].height = 20
r += 1

notas = [
    "• Los totales de cada apartado se calculan sin filtrar por STATUS (suma completa del bloque).",
    "• Top 5 por VENDEDOR: Cobranza Corriente, Cancelaciones.",
    "• Top 5 por UBICACIÓN: Cobranza Vencida.",
    "• Top 5 por COBRADO: Cobranza Efectiva, Recuperada, Adelantada F., Total General.",
    "• Los montos de TRANSFERENCIAS / DEPÓSITOS se muestran consolidados como 'TRANSFER. / DEPTOS.' en Top 5.",
    "• Proyección = estimación de cobranza para el mes siguiente basada en tendencia.",
    "• Colores: Encabezados azul oscuro (#1F3C6E) | Acentos dorado (#C9A84C) | Filas alternadas.",
]
for nota in notas:
    cell_style(ws, r, 1, nota, bg=BLANCO, fg=GRIS_TEXTO, size=9,
               align="left")
    ws.merge_cells(start_row=r, start_column=1,
                  end_row=r, end_column=COLS)
    ws.row_dimensions[r].height = 16
    r += 1

# ══ 7) PRINT SETUP ═════════════════════════════════════════
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_setup.fitToPage = True
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.print_title_rows = "1:3"

# ══ GUARDADO ════════════════════════════════════════════════
out_dir = "/home/elena/.openclaw/workspace/workspaces/cobranza/estadisticas"
os.makedirs(out_dir, exist_ok=True)
path = f"{out_dir}/PLANTILLA_ESTADISTICA_JUNTA_MENSUAL_2026.xlsx"
wb.save(path)
print(f"✅ Plantilla guardada: {path}")
print(f"   Filas usadas: ~{r}  |  Columnas: {COLS}")