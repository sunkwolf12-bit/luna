"""
Carga datos reales de JULIO 2025 en la plantilla 2026 (ENERO como mes muestra)
para que Elena pueda evaluar el formato.
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.cell.cell import MergedCell

AZUL_CLARO     = "D6E4F0"
DORADO_CLARO   = "F5E6B8"
BLANCO         = "FFFFFF"
GRIS_CLARO     = "F2F2F2"
GRIS_TEXTO     = "404040"
NEGRO          = "1A1A1A"

def mkfill(ws, row, col, value=None, bg=BLANCO, fg=GRIS_TEXTO,
           bold=False, size=10, align="center", fmt=None):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return  # skip merged cells
    cell.value = value
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Calibri", size=size, bold=bold, color=fg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        cell.number_format = fmt

wb = load_workbook(
    "/home/elena/.openclaw/workspace/workspaces/cobranza/estadisticas/"
    "PLANTILLA_ESTADISTICA_JUNTA_MENSUAL_2026.xlsx"
)
ws = wb["ESTADISTICA 2026"]

# ── DATOS DE JULIO 2025 (ya validados) ───────────────────
JULIO = {
    "corriente":     1_155_493,
    "cancelaciones":   122_087,
    "vencida":         188_087,
    "efectiva":       1_157_835,
    "recuperada":       141_683,
    "total_gral":     1_363_176,
    "adelantada":        16_280,
    "proyeccion":     1_063_756,
}

TOP5 = {
    "corriente": [
        ("V6",  289_000), ("V1", 250_000),
        ("V39", 238_000), ("V38", 200_000), ("V27", 122_000),
    ],
    "cancelaciones": [
        ("V6",  22_300), ("V84", 16_400), ("V38", 15_900),
        ("V56", 15_700), ("V55", 15_600),
    ],
    "vencida": [
        ("V39", 30_200), ("EDUARDO", 27_900), ("DEPOSITO", 27_300),
        ("FRANCISCO", 25_800), ("JORGE", 21_200),
    ],
    "efectiva": [
        ("FRANCISCO", 168_000), ("OFICINA", 127_000),
        ("EDGAR", 127_000), ("DEPOSITO", 118_000), ("EDUARDO", 117_000),
    ],
    "recuperada": [
        ("EDUARDO", 26_100), ("V39", 21_400), ("FRANCISCO", 20_300),
        ("JORGE", 17_500), ("EDGAR", 12_300),
    ],
    "total_gral": [
        ("FRANCISCO", 194_000), ("DEPOSITO", 152_000),
        ("EDUARDO", 146_000), ("OFICINA", 142_000), ("EDGAR", 141_000),
    ],
    "adelantada": [("OFICINA", 12_000), ("V6", 4_280)],
}

# ── Estructura conocida de ENERO 2026 (fila 5) ────────────
# Fila 5: título "ENERO 2026" (merged 1-14)
# Fila 6: encabezados
# Fila 7: "DATOS GENERALES" section header
# Fila 8: Mes | [valor merge B-D] | Año | [valor merge E-H]
# Fila 9: Fecha de corte | [valor merge B-D] | Elaboró | [valor merge E-H]
# Fila 10: Reportado por | [valor merge B-D] | Cargo | [valor merge E-H]
# Fila 11: empty
# Fila 12: COBRANZA CORRIENTE section header
# Fila 13: Total | [valor en B]
# Fila 14: Top 5 | [merge B-C]
# Fila 15-19: 1er-5o lugar | nombre | monto | %TOP1
# Fila 20: empty
# ... same pattern for each category ...

def fill_apartado(r, bg_hdr, total, top5_list):
    """Llena un apartado completo. Devuelve la siguiente fila libre."""
    # Fila Total
    mkfill(ws, r, 1, "Total", bg=bg_hdr, fg=GRIS_TEXTO, bold=True, align="left")
    mkfill(ws, r, 2, total, bg=bg_hdr, fg=NEGRO, bold=True, size=11, align="center",
           fmt='"$"#,##0.00')
    r += 1
    # Fila Top 5 header
    mkfill(ws, r, 1, "Top 5", bg=DORADO_CLARO, fg=GRIS_TEXTO, bold=True, size=9)
    r += 1
    # Filas 1-5
    for i, (nom, mon) in enumerate(top5_list, 1):
        mkfill(ws, r, 1, f"{i}er lugar" if i == 1 else f"{i}o lugar",
               bg=BLANCO, fg=GRIS_TEXTO, size=9, align="left")
        mkfill(ws, r, 2, nom, bg=BLANCO, fg=GRIS_TEXTO, size=9, align="center")
        mkfill(ws, r, 3, mon, bg=BLANCO, fg=NEGRO, size=10, align="center",
               fmt='"$"#,##0.00')
        pct = mon / total if total else 0
        mkfill(ws, r, 8, pct, bg=BLANCO, fg=GRIS_TEXTO, size=9, align="center",
               fmt='0.0%')
        r += 1
    r += 1  # separador
    return r

# ── DATOS GENERALES ─────────────────────────────────────
mkfill(ws, 8, 1, "Mes", bg=AZUL_CLARO, fg=GRIS_TEXTO, bold=True, align="left")
mkfill(ws, 8, 2, "JULIO 2025 — MUESTRA", bg=AZUL_CLARO, fg=NEGRO, bold=True,
       size=10, align="left")
# Año
mkfill(ws, 9, 1, "Fecha de corte", bg=AZUL_CLARO, fg=GRIS_TEXTO, bold=True, align="left")
mkfill(ws, 9, 2, "31/07/2025", bg=AZUL_CLARO, fg=GRIS_TEXTO, size=10, align="left")
mkfill(ws, 10, 1, "Elaboró", bg=AZUL_CLARO, fg=GRIS_TEXTO, bold=True, align="left")
mkfill(ws, 10, 2, "Elena Rivas", bg=AZUL_CLARO, fg=GRIS_TEXTO, size=10, align="left")

# ── CATEGORÍAS ──────────────────────────────────────────
cats = [
    ("corriente",     "COBRANZA CORRIENTE",    AZUL_CLARO),
    ("cancelaciones", "CANCELACIONES",          GRIS_CLARO),
    ("vencida",       "COBRANZA VENCIDA",      AZUL_CLARO),
    ("efectiva",      "COBRANZA EFECTIVA",     GRIS_CLARO),
    ("recuperada",    "COBRANZA RECUPERADA",   GRIS_CLARO),
    ("total_gral",    "COBRANZA TOTAL GENERAL", GRIS_CLARO),
    ("adelantada",    "COBRANZA ADELANTADA F.",AZUL_CLARO),
]

r = 13  # primera fila de COBRANZA CORRIENTE
for key, nombre, bg in cats:
    r = fill_apartado(r, bg, JULIO[key], TOP5[key])

# ── PROYECCIÓN ──────────────────────────────────────────
# r ya apunta a la fila de PROYECCIÓN tras el último apartado
mkfill(ws, r, 1, f"Proyección AGOSTO 2025", bg=DORADO_CLARO, fg=GRIS_TEXTO,
       bold=True, align="left")
mkfill(ws, r, 2, JULIO["proyeccion"], bg=DORADO_CLARO, fg=NEGRO, bold=True,
       size=11, align="center", fmt='"$"#,##0.00')

out = ("/home/elena/.openclaw/workspace/workspaces/cobranza/estadisticas/"
       "PLANTILLA_ESTADISTICA_JUNTA_MENSUAL_2026_CON_DATOS.xlsx")
wb.save(out)
print(f"✅ Guardado: {out}")
print(f"   Última fila escrita: ~{r}")