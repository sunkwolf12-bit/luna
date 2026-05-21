"""
Genera la PLANTILLA CORREGIDA + CARGA ENERO 2026
Columnas: A=Concepto | B=TOTAL | C=Top1 | D=Monto1 | E=Top2 | F=Monto2 ...
y carga todos los datos de ENERO 2026.
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os

AZUL_OBSCURO   = "1F3C6E"
AZUL_MEDIO     = "2E5090"
AZUL_CLARO     = "D6E4F0"
DORADO         = "C9A84C"
DORADO_CLARO   = "F5E6B8"
BLANCO         = "FFFFFF"
GRIS_CLARO     = "F2F2F2"
GRIS_TEXTO     = "404040"
NEGRO          = "1A1A1A"
VERDE          = "1A5C1F"  # para % Top1 destacado

# ── helpers ────────────────────────────────────────────────
def fill(ws, row, col, value=None, bg=BLANCO, fg=GRIS_TEXTO,
         bold=False, size=10, align="center", fmt=None, italic=False):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.fill = PatternFill("solid", fgColor=bg)
    font_kwargs = dict(name="Calibri", size=size, bold=bold,
                       color=fg, italic=italic)
    cell.font = Font(**font_kwargs)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=False)
    if fmt:
        cell.number_format = fmt

def section_bar(ws, row, text, cols, color=AZUL_MEDIO):
    fill(ws, row, 1, text.upper(), bg=color,
         fg=DORADO, bold=True, size=10, align="left")
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=cols)

# ── DATOS ENERO 2026 ──────────────────────────────────────
T = {
    "corriente":    1_608_391,
    "cancelaciones":  127_110,
    "vencida":        166_220,
    "efectiva":      1_228_968,
    "total_gral":    1_408_644,
    "recuperada":      115_277,
    "adelantada":       80_464,
    "proyeccion":    1_158_184,
}

# Top5: (nombre, monto)
TOP5 = {
    "corriente": [
        ("V6",  356_324), ("V1",  297_307), ("V38", 241_116),
        ("V39", 176_093), ("V56", 130_775),
    ],
    "cancelaciones": [
        ("V38", 21_035), ("V56", 17_664), ("V6", 15_924),
        ("V84", 14_374), ("V1",  13_188),
    ],
    "vencida": [
        ("V39", 32_232), ("FRANCISCO", 24_988), ("JORGE", 23_648),
        ("DEPOSITO", 19_599), ("EDGAR", 16_160),
    ],
    "efectiva": [
        ("EDGAR", 151_234), ("FRANCISCO", 141_345),
        ("EDUARDO", 91_734), ("JORGE", 88_734), ("OFICINA", 27_924),
    ],
    "total_gral": [
        ("OFICINA", 127_234), ("EDGAR", 127_000),
        ("FRANCISCO", 122_000), ("DEPOSITO", 110_000), ("EDUARDO", 106_000),
    ],
    "recuperada": [
        ("EDUARDO", 26_100), ("V39", 21_400), ("FRANCISCO", 20_300),
        ("JORGE", 17_500), ("EDGAR", 12_300),
    ],
    "adelantada": [
        ("OFICINA", 49_405), ("V6", 31_059),
    ],
}

MESES_2026 = [
    ("ENERO",     "01"), ("FEBRERO",   "02"), ("MARZO",     "03"),
    ("ABRIL",     "04"), ("MAYO",      "05"), ("JUNIO",     "06"),
    ("JULIO",     "07"), ("AGOSTO",    "08"), ("SEPTIEMBRE","09"),
    ("OCTUBRE",   "10"), ("NOVIEMBRE", "11"), ("DICIEMBRE", "12"),
]

# Layout de columnas:
# A=Concepto | B=TOTAL | C=Vendedor1 | D=Monto1 | E=Vendedor2 | F=Monto2
# G=Vendedor3 | H=Monto3 | I=Vendedor4 | J=Monto4 | K=Vendedor5 | L=Monto5
# M=%Top1 | N=%Acum
COLS = 14

# ── CONSTRUIR LIBRO ─────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "ESTADISTICA 2026"

# Anchos de columna
ws.column_dimensions["A"].width = 28
for c in range(2, COLS + 1):
    ws.column_dimensions[get_column_letter(c)].width = 13
ws.freeze_panes = "B4"

# ── ENCABEZADO MAESTRO ──────────────────────────────────
fill(ws, 1, 1, "PROTEG-RT MUTUALIDAD", bg=AZUL_OBSCURO, fg=DORADO,
     bold=True, size=16, align="center")
ws.merge_cells("A1:N1")
fill(ws, 2, 1, "REPORTE DE COBRANZA — JUNTA MENSUAL  |  AÑO 2026",
     bg=AZUL_MEDIO, fg=BLANCO, bold=True, size=12, align="center")
ws.merge_cells("A2:N2")
fill(ws, 3, 1, "Plantilla corregida — Mayo 2026",
     bg=AZUL_MEDIO, fg=DORADO, bold=False, size=10, align="center")
ws.merge_cells("A3:N3")
ws.row_dimensions[1].height = 26
ws.row_dimensions[2].height = 20
ws.row_dimensions[3].height = 16

# ── SUB-ENCABEZADOS DE COLUMNA ──────────────────────────
col_headers = [
    "CONCEPTO",          # A
    "TOTAL ($)",         # B
    "TOP 1 — Vendedor",  # C
    "TOP 1 — Monto",     # D
    "TOP 2 — Vendedor",  # E
    "TOP 2 — Monto",     # F
    "TOP 3 — Vendedor",  # G
    "TOP 3 — Monto",     # H
    "TOP 4 — Vendedor",  # I
    "TOP 4 — Monto",     # J
    "TOP 5 — Vendedor",  # K
    "TOP 5 — Monto",     # L
    "% Top 1",           # M
    "% Acum.",           # N
]
for c, h in enumerate(col_headers, 1):
    fill(ws, 4, c, h, bg=DORADO, fg=NEGRO, bold=True, size=9, align="center")
ws.row_dimensions[4].height = 30

# ── helper: mes block ──────────────────────────────────
def build_month_block(ws, start_row, mes_nombre, mes_num,
                      data, top5, bg_hdr):
    r = start_row
    ws.row_dimensions[r].height = 22

    # Barra de mes
    fill(ws, r, 1, f"{mes_nombre} 2026", bg=AZUL_OBSCURO,
         fg=DORADO, bold=True, size=13, align="center")
    ws.merge_cells(start_row=r, start_column=1,
                   end_row=r, end_column=COLS)
    r += 1

    # Helper: apartado
    def apartado(label, totalkey, top5key, bg_row):
        nonlocal r
        section_bar(ws, r, label, COLS)
        ws.row_dimensions[r].height = 18
        r += 1
        # Fila TOTAL
        fill(ws, r, 1, "  Total", bg=bg_row, fg=GRIS_TEXTO,
             bold=True, size=10, align="left")
        fill(ws, r, 2, data[totalkey], bg=bg_row, fg=NEGRO,
             bold=True, size=12, align="center",
             fmt='"$"#,##0.00')
        ws.row_dimensions[r].height = 20
        r += 1
        # Fila Top 5
        fill(ws, r, 1, "  Top 5", bg=DORADO_CLARO, fg=GRIS_TEXTO,
             bold=True, size=9, align="left")
        ws.row_dimensions[r].height = 16
        r += 1
        # Filas 1-5
        rows_top5 = top5[top5key]
        running = 0
        for i, (nom, mon) in enumerate(rows_top5, 1):
            lbl = f"  {i}er lugar" if i == 1 else f"  {i}o lugar"
            fill(ws, r, 1, lbl, bg=BLANCO, fg=GRIS_TEXTO,
                 size=9, align="left")
            # vendedor col (C, E, G, I, K)
            vcol = 3 + (i - 1) * 2
            mcol = vcol + 1
            fill(ws, r, vcol, nom, bg=BLANCO, fg=GRIS_TEXTO,
                 size=10, align="center")
            fill(ws, r, mcol, mon, bg=BLANCO, fg=NEGRO,
                 size=10, align="center", fmt='"$"#,##0.00')
            # %Top1
            pct1 = mon / data[totalkey] if data[totalkey] else 0
            fill(ws, r, 13, pct1, bg=BLANCO, fg=GRIS_TEXTO,
                 size=9, align="center", fmt='0.0%')
            # %Acum
            running += mon
            pct_acum = running / data[totalkey] if data[totalkey] else 0
            fill(ws, r, 14, pct_acum, bg=BLANCO, fg=GRIS_TEXTO,
                 size=9, align="center", fmt='0.0%')
            ws.row_dimensions[r].height = 17
            r += 1
        # Vacío separador
        ws.row_dimensions[r].height = 6
        r += 1

    apartado("COBRANZA CORRIENTE",     "corriente",    "corriente",    AZUL_CLARO)
    apartado("CANCELACIONES",          "cancelaciones","cancelaciones",GRIS_CLARO)
    apartado("COBRANZA VENCIDA",        "vencida",      "vencida",      AZUL_CLARO)
    apartado("COBRANZA EFECTIVA",       "efectiva",     "efectiva",     GRIS_CLARO)
    apartado("COBRANZA RECUPERADA",     "recuperada",   "recuperada",   GRIS_CLARO)
    apartado("COBRANZA TOTAL GENERAL",  "total_gral",   "total_gral",   GRIS_CLARO)
    apartado("COBRANZA ADELANTADA F.",  "adelantada",   "adelantada",   AZUL_CLARO)

    # Proyección
    section_bar(ws, r, f"PROYECCIÓN {mes_nombre} → {mes_num}",
                COLS, color=DORADO)
    ws.row_dimensions[r].height = 18
    r += 1
    fill(ws, r, 1, f"  Proyección {mes_num}",
         bg=DORADO_CLARO, fg=GRIS_TEXTO, bold=True, size=10, align="left")
    fill(ws, r, 2, data["proyeccion"], bg=DORADO_CLARO, fg=NEGRO,
         bold=True, size=12, align="center", fmt='"$"#,##0.00')
    ws.row_dimensions[r].height = 20
    r += 1
    ws.row_dimensions[r].height = 12
    r += 1

    return r

# ── ENERO 2026 (con datos reales) ───────────────────────
r = build_month_block(ws, 5, "ENERO", "FEBRERO 2026",
                      T, TOP5, AZUL_CLARO)

# ── RESTO DE MESES (vacíos) ─────────────────────────────
ws.row_dimensions[r].height = 14
r += 1
for mes_nombre, mes_num in MESES_2026[1:]:
    ws.row_dimensions[r].height = 22
    fill(ws, r, 1, f"{mes_nombre} 2026", bg=AZUL_OBSCURO,
         fg=DORADO, bold=True, size=13, align="center")
    ws.merge_cells(start_row=r, start_column=1,
                   end_row=r, end_column=COLS)
    r += 1
    fill(ws, r, 1, "  Total — pendiente", bg=GRIS_CLARO,
         fg=GRIS_TEXTO, size=10, align="left")
    ws.row_dimensions[r].height = 20
    r += 1
    ws.row_dimensions[r].height = 10
    r += 1

# ── NOTAS ────────────────────────────────────────────────
r += 1
fill(ws, r, 1, "NOTAS Y LEYENDA", bg=AZUL_OBSCURO, fg=DORADO,
     bold=True, size=10, align="center")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
ws.row_dimensions[r].height = 18
r += 1

notas = [
    "• Los totales de cada apartado se calculan sin filtrar por STATUS (suma completa del bloque).",
    "• Top 5 por VENDEDOR: Cobranza Corriente, Cancelaciones.",
    "• Top 5 por UBICACIÓN: Cobranza Vencida.",
    "• Top 5 por COBRADO: Cobranza Efectiva, Recuperada, Adelantada F., Total General.",
    "• Los montos de TRANSFERENCIAS / DEPÓSITOS se muestran consolidados como 'TRANSFER. / DEPTOS.' en Top 5.",
    "• Proyección = estimación de cobranza para el mes siguiente basada en tendencia.",
    "• Colores: Encabezados azul oscuro (#1F3C6E) | Acentos dorado (#C9A84C) | Filas alternadas.",
]
for nota in notas:
    fill(ws, r, 1, nota, bg=BLANCO, fg=GRIS_TEXTO, size=9, align="left")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=COLS)
    ws.row_dimensions[r].height = 15
    r += 1

# ── PRINT SETUP ─────────────────────────────────────────
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToPage = True
ws.print_title_rows = "1:4"

# ── GUARDAR ─────────────────────────────────────────────
out_dir = "/home/elena/.openclaw/workspace/workspaces/cobranza/estadisticas"
os.makedirs(out_dir, exist_ok=True)
path = f"{out_dir}/ESTADISTICA_JUNTA_MENSUAL_2026_ENERO.xlsx"
wb.save(path)
print(f"✅ Guardado: {path}")
print(f"   Filas usadas: ~{r}")